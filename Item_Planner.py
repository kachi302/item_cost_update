import sys
import os

import pyautogui as pag
import time
import datetime as dt
from openpyxl import load_workbook
import pymsgbox as pmg

running = True
print('Item Planner Code 변경...3초후 시작됩니다. 해당 화면을 다른 모니터로 이동 하세요.')
pag.countdown(3)
# item_code = ''
# planner_code = ''18669305-004
try:
    print('Start')
    # pag.click(1000, 800)
    wb = load_workbook(r'D:\python_dev\item_cost_update\planner.xlsx', data_only= True)
    ws = wb['Sheet1']
    start1 = time.time()
    rec = 0
    time.sleep(1)
    dest = "D:\\python_dev\\item_cost_update\\images\\"
    # job = pag..confirm('Item Planner 변경 작업을 진행 하시겠습니끼?', buttons= ['OK','Cancel']) # type: ignore    
    job = pmg.confirm('Item Planner 변경 작업을 진행 하시겠습니끼?', buttons= ['OK','Cancel']) 
    if job == 'OK':
        print(time.strftime('%Y.%m%d - %H:%M:%S'))
        time.sleep(3)
        for row in ws.iter_rows():
            rec += 1
            item = str(row[0].value)
            time.sleep(0.3)
            item_len = len(item)
            
            
            if item =='E':
                print('End....')
                break
            if item_len <= 11:
                print('Rec : %d , item = %s, %s' % (rec,item, 'item error'))
                break
            
            planner = str(row[1].value)
            time.sleep(0.5)
            pag.hotkey('f11')
            time.sleep(1)
            pag.write(item, interval=0.03)
            time.sleep(0.5)
            pag.hotkey('ctrl','f11')
            time.sleep(1)
            main_png = pag.locateAllOnScreen(r'D:\python_dev\item_cost_update\images\main_01.png', confidence=0.8)
            time.sleep(0.5)
            if main_png is None:
                main_png2 = pag.locateOnScreen(r'D:\python_dev\item_cost_update\images\item_menu02.png', confidence=0.8)
                if main_png2 is not None:
                    pag.click(pag.center(main_png2))
                    time.sleep(0.5)
                main_png02=pag.locateOnScreen(r'D:\python_dev\item_cost_update\images\main_02.png', confidence=0.8)
                if main_png02 is not None:
                    pag.click(pag.center(main_png02))
                    time.sleep(0.5)
                
            menu_png = pag.locateOnScreen(r'D:\python_dev\item_cost_update\images\item_menu01.png', confidence=0.8)
            time.sleep(1)
            if menu_png is not None:
                pag.click(pag.center(menu_png))
                gp_menu_png = pag.locateOnScreen(r'D:\python_dev\item_cost_update\images\gl_menu.png', confidence=0.8)
                # pag.press('down', presses=8)
                if gp_menu_png is not None:
                    pag.click(pag.center(gp_menu_png))
                    time.sleep(0.5)
                    
                # pag.press('enter')
                pag.press('tab')
                time.sleep(0.5)
                pag.write(planner, interval=0.2)
                time.sleep(0.5)
                pag.hotkey('ctrl','s')
                time.sleep(2)
                menu_png1 = pag.locateOnScreen(r'D:\python_dev\item_cost_update\images\item_menu02.png', confidence=0.8)
                time.sleep(0.5)
                if menu_png1 is not None:
                    pag.click(pag.center(menu_png1))
                    pag.press('up', presses=8)
                    time.sleep(0.5)
                    pag.press('enter')
                    time.sleep(1)
                else:
                    print('Rec : %d , item = %s, %s' % (rec,item, 'menu2 error'))
                    pag.screenshot(dest+"menu2_second.jpg")
                    time.sleep(1)
                    break
            else:
                print('Rec : %d , item = %s, %s' % (rec,item, 'menu1 error'))
                pag.screenshot(dest+"menu1_first.jpg")
                time.sleep(1)
                break   
            print('Rec : %d , item = %s' % (rec,item))
    print(time.strftime('%Y.%m%d - %H:%M:%S'))        
    pmg.alert('작업을 종료합니다.')
    print('작업을 종료합니다.')
except Exception as e:
    # wb.active = ws # type: ignore
    # wb.save(r'.\Planner_job.xlsx')    # type: ignore
    print(e)    
    time.sleep(10)       