
import sys
import cgitb
from cgitb import text
import glob
cgitb.enable()

from tkinter import messagebox, Tk

import tkinter as tk

import pyautogui
import time
from openpyxl import load_workbook
import datetime as dt
import configparser

config = configparser.ConfigParser()
config.read(r'.\data\config.ini')

ct_x1 = config.getint('Screen_3WAY', 'X1')
ct_x2 = config.getint('Screen_3WAY', 'X2')
ct_x3 = config.getint('Screen_3WAY', 'X3')
ct_x4 = config.getint('Screen_3WAY', 'X4')

dx1 = config.getint('Screen_Not_Default', 'DX1')
dy1 = config.getint('Screen_Not_Default', 'DY1')
dx2 = config.getint('Screen_Not_Default', 'DX2')
dy2 = config.getint('Screen_Not_Default', 'DY2')

running = True

def on_escape(event):
    mainFrame.destroy()    
    
def update_progress(new_progress):
    progress_label.config(text=new_progress)
    mainFrame.update()

# def count_total_records(file):
#     # wb = load_workbook(r'.\data\ANNUAL_list.xlsx')
#     wb = load_workbook(file)
#     ws = wb['Sheet1']
#     total_records = sum(1 for _ in ws.iter_rows()) - 1  # 엑셀 헤더 행을 제외한 전체 레코드 수 계산
#     return total_records

def cost_menu_image_check(sel_image, sel_item, sel_cost_type):
    try:
        check_image = pyautogui.locateOnScreen(sel_image)
        if check_image is not None:
            center=pyautogui.center(check_image)
            pyautogui.doubleClick(center)
            time.sleep(0.8)
            pyautogui.write(sel_item, interval=0.1)
            time.sleep(0.8)
            pyautogui.hotkey('tab')
            time.sleep(0.8)
            # cost type 입력
            pyautogui.write(sel_cost_type, interval=0.1)  
            time.sleep(0.8)
            pyautogui.hotkey('tab')  
            time.sleep(0.8)
            fi = pyautogui.locateOnScreen(r'.\images\find.png')
            fic = pyautogui.center(fi) # type: ignore
            pyautogui.click(fic)                    
            time.sleep(1)      
            return 'Y'
        else:
            # return 'NO'
            return 'N'
    except:
        messagebox.showerror("Error", "Cost Menu , 해당 ITEM 이 존재하지 않습니다. 오류가 발생했습니다.")
        return 'N'
    
    
def annual_reset():
    try:
        label_result.config(text='Annual Cost Reset')
        wb = load_workbook(r'.\data\ANNUAL_list.xlsx')
        # wb = load_workbook(r'C:\ITEM_COST_UPDATE\data\ANNUAL_list.xlsx')
        ws = wb['Sheet1']
        # total_records = count_total_records(u_file)  # 전체 레코드 수 계산
        total_records = ws.max_row
        
        rec = 0
        time.sleep(3)
        running = True
        job = pyautogui.confirm(text='Annual Cost Reset 작업을 진행 하시겠습니까 ?', buttons=['OK', 'Cancel']) # type: ignore
        progress_label.config(text='진행 중...')
        mainFrame.update()

        if job == 'OK':
            for idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
                if running == False:
                    break

                cell_values = list(row)
                if not all(cell_values):  # 모든 셀 값이 비어 있는 경우 작업 중지
                    running = False
                    break

                item = cell_values[0]
                
                update_progress(f'진행 상황: {idx}/{total_records} : {item}')  # 현재 진행 상황 업데이트
                
                check_menu_YN= cost_menu_image_check(item_mp, item, 'Annual')
                if check_menu_YN =='Y':
                    # menu 가 있는 경우
                    r_check1 = cost_type_3way(cost_3_yes)
                    if r_check1 == 'N':

                        # default control 이 아닌 경우, inventory , cost roll up 인 경우
                        r_check2 = cost_type_not_default(cost_not_default_img)
                        
                        if r_check2 == 'N':
                            r_check3 = cost_type_inventory(cost_type_inventory_img, 0, 0)
                            #inventory asset 이 아닌 경우
                            if r_check3 == 'N':
                                pyautogui.hotkey('ctrl', 'f4')
                                time.sleep(1)
                                pyautogui.hotkey('ctrl', 'f4')
                                time.sleep(1)
                                running = False
                                break

                    
                else:
                    update_progress(f'해당 Item의 Annual Type 여부 확인 필요, Error : {idx}/{total_records}{item}')  # 현재 진행 상황 업데이트    
                    break
            messagebox.showinfo('Annual Reset', '작업 완료')    
        else:
            return 0

    except:
        messagebox.showerror('Error', ' Annual Reset, 오류가 발생 했습니다.')
        
        
# Max screen
def max_screen():
    try:
        pyautogui.click(1000, 800)
        screen_max = pyautogui.locateOnScreen(r'.\images\screen_max.png',confidence=0.8)
        if screen_max != None:
            screen_max_ce = pyautogui.center(screen_max)
        
            pyautogui.click(screen_max_ce.x , screen_max_ce.y-20, interval=0.5)
            messagebox.showinfo('Max Screen', 'ERP 화면을 Max 로 ')
        else:
            messagebox.showinfo('Max Screen', 'ERP 화면을 Max 로 하세요.')
    except:
        messagebox.showerror('Error',' ERP 초기 화면 상태 설정 Error')    

def item_cost_type_not_found():
    try:
        type_not_found = pyautogui.locateOnScreen(r'.\images\cost_type_not_found.png')
        if type_not_found != None:
            return 'Y'
        else:
            return 'N'                
    except:
        messagebox.showerror('Error',' Cost Type not found !!!')    
        return 'N'                    
                     
def item_new_cost(item, cost_type):
    try:
        new_type = pyautogui.locateOnScreen(r'.\images\new.png')
        if new_type  != None:
            new_center = pyautogui.center(new_type)
            pyautogui.click(new_center)
            time.sleep(1)
            # item code d
            pyautogui.write(item, interval=0.1)
            time.sleep(0.5)
            pyautogui.hotkey('tab')
            pyautogui.write(cost_type, interval=0.1)  
            time.sleep(0.5)
            # new 일경우
            # cost_type_inventory(cost_type_inventory_img)
            return 'Y'
    except:
        messagebox.showerror('Error',' New !!!')    
        return 'N'            
                     
def annual_update():
    try:
        global cost_element1
        global cost_element2
        global sub_element2
        
        max_screen()
        label_result.config(text='Annual Cost Update...')
        u_file = r'.\data\ANNUAL_UPDATE.xlsx'
        wb = load_workbook(r'.\data\ANNUAL_UPDATE.xlsx')
        ws = wb['Sheet1']
        # total_records = count_total_records(u_file)  # 전체 레
        total_records = ws.max_row
        job = pyautogui.confirm(text='Annual Cost Update 작업을 진행 하시겠습니까 ?', buttons=['OK', 'Cancel'])   # type: ignore
        progress_label.config(text='Annual Cost Update 진행 중...')
        mainFrame.update()
        
        if job == 'OK':
            rec = 0
            time.sleep(3)
          
            for row in ws.iter_rows():
                rec += 1
                
                item = str(row[0].value)
                cost_type = str(row[1].value)
                cost = str(row[2].value)
                overhead = str(row[3].value)
                
                update_progress(f'Annual Update 진행 상황: {rec}/{total_records} , {item}')  # 현재 진행 상황 업데이트
                # 
                cost_menu_image_check(item_mp, item, cost_type)
                # 해당 cost type 이 존재 하는 경우
                check_not_cost = item_cost_type_not_found()  
                # 해당 cost type 이 존재 여부 확인
                if check_not_cost == 'Y':
                    #해당 cost type 이 존재 하지 않는다.
                    check_new = item_new_cost(item, cost_type)
                    if check_new =='Y':
                    # new 이면서 inventory asset 인경우
                        cost_type_inventory(cost_type_inventory_img,cost, overhead )
                        
                else:
                    # cost type 이 존해 경우
                    # messagebox.showerror("Error", " N, COST 항목 full check 부분, No 발생했습니다.")
                    # cost type 이 존재 하는 경우, inventory asset 인지 확인 한다.
                    check_inv_YN = cost_type_inventory(cost_type_inventory_img,cost, overhead )
                    if check_inv_YN == 'N':
                        update_progress(f'Annual Update 중 Error : {rec}/{total_records} , {item}') 
                        break
            messagebox.showinfo('Annaul Update', 'Annual Cost update 완료')
    except:
        messagebox.showerror('Error',' Annual Cost Update error') 
        
def cost_type_3way(cost_3_yes):
    try:
        global ct_x1
        global ct_x2
        global ct_x3
        global ct_x4
        global dx1
        global dx2
        global dy1
        global dy2
        
        check_3way = pyautogui.locateOnScreen(cost_3_yes)
        if check_3way != None:
            ric = pyautogui.center(check_3way) 
            # user defaul control uncheck
            time.sleep(0.7)
            # pyautogui.click(ric.x-130, ric.y)
            pyautogui.click(ric.x+ct_x1, ric.y)
                            
                            # based on rollup unckeck
            time.sleep(0.7)
            # pyautogui.click(ric.x+130, ric.y)
            pyautogui.click(ric.x+ct_x2, ric.y)
            time.sleep(1)
                            # cost 지울지 메세지 나타나남, enter key 입력
            pyautogui.hotkey('enter')
            time.sleep(1)                                                 
                            # based on rollup ckeck
            # pyautogui.click(ric.x+130, ric.y)
            pyautogui.click(ric.x+ct_x3, ric.y)                       
            time.sleep(1)  
                            
                            # user defaul control check
            # pyautogui.click(ric.x-130, ric.y)
            pyautogui.click(ric.x+ct_x4, ric.y)
            time.sleep(1)
                            
            pyautogui.hotkey('ctrl','s')
            time.sleep(1)
            pyautogui.hotkey('ctrl','f4')
                        # close_form()    
            time.sleep(1)
            return 'Y'
        else:
            return 'N'
    except:
        messagebox.showerror("Error", " COST 항목 full check 부분, 오류가 발생했습니다.")


def  cost_type_not_default(cost_not_default_png):
    # default control 이 disable 이고 inventory asset 과 baseed on rollup 은 check 되어 있는 경우 
    # Annual Cost reset process
    try:
        global ct_x1
        global ct_x2
        global ct_x3
        global ct_x4
        global dx1
        global dx2
        global dy1
        global dy2
        
        check_not_default =   pyautogui.locateOnScreen(cost_not_default_png) 
        if check_not_default != None:
        
            ri_undefault_c = pyautogui.center(check_not_default)                            
            time.sleep(0.9)  
            
            # pyautogui.click(ri_undefault_c.x+115, ri_undefault_c.y+10)
            pyautogui.click(ri_undefault_c.x+dx1, ri_undefault_c.y+dy1)
            time.sleep(0.9)
            
            # cost 지울지 메세지 나타나남, enter key 입력
            pyautogui.hotkey('enter')
            time.sleep(0.9)
            
            # pyautogui.click(ri_undefault_c.x+118, ri_undefault_c.y+10)
            pyautogui.click(ri_undefault_c.x+dx2, ri_undefault_c.y+dy2)
            time.sleep(0.9)
            pyautogui.hotkey('ctrl','s')
            # close_form()   
            time.sleep(1)
            pyautogui.hotkey('ctrl','f4')               
            time.sleep(1)
            return 'Y'  
        else:
            return 'N'
    except:
        messagebox.showerror("Error", "Default Control is disabled, 오류가 발생했습니다.")
            
def cost_type_inventory(cost_type_inventory_img, cost_value, overhead_rate):
    try:
        inv_asset = pyautogui.locateOnScreen(cost_type_inventory_img)
        if  inv_asset != None :     
            check_cost_YN = cost_update_function(cost_value, overhead_rate)
            if check_cost_YN == 'Y':
                return 'Y'
            else:
                return 'N'
        else:
            return 'N'
    except:
        messagebox.showerror("Error", "Inventory Asset update, 오류가 발생했습니다.")
        return 'N'

def cost_update_function(cost_value, overhead_rate):
    try:
        global cost_element1
        global cost_element2
        global sub_element2 
        
        time.sleep(0.9)
        # cost_btn = pyautogui.locateOnScreen(resource_path(r'images\costs_btn.png'))
        cost_btn = pyautogui.locateOnScreen(r'.\images\costs_btn.png')
        cost_btn_center=pyautogui.center(cost_btn) # type: ignore
        pyautogui.click(cost_btn_center)
        time.sleep(1.5)                               
        
        pyautogui.write(cost_element1, interval=0.1)
        time.sleep(0.5)
        
        pyautogui.hotkey('tab')
        pyautogui.write(cost_element1, interval=0.1)
        time.sleep(0.5)           
                    # activity 칼럼 이동                  
        pyautogui.press('tab')
        time.sleep(0.5)     
        # basis 칼럼                   
        pyautogui.press('tab')
        time.sleep(0.5)
        
        pyautogui.write('Item', interval=0.1) 
        time.sleep(0.5)
        pyautogui.hotkey('tab')
        time.sleep(0.5)
        # Material cost 입력
        # pyautogui.write(str(0), interval=0.1)
        pyautogui.write(str(cost_value), interval=0.1)
        time.sleep(0.9)
        pyautogui.hotkey('tab')
        time.sleep(0.9)
        pyautogui.hotkey('tab')
        # overhead
        pyautogui.write(cost_element2, interval=0.1)
        time.sleep(0.5)
        pyautogui.hotkey('tab')
        # sub element
        pyautogui.write(sub_element2, interval=0.1)
        time.sleep(0.6)
        pyautogui.hotkey('tab')
        time.sleep(0.5)
        pyautogui.press('tab') 
        time.sleep(0.5)
        pyautogui.write('Total Value', interval=0.1)
        # Material overhead rate 입력         
        pyautogui.hotkey('tab')
        time.sleep(0.6)
        pyautogui.write(str(overhead_rate), interval=0.1)
        time.sleep(0.6)
        pyautogui.hotkey('tab')                        
        time.sleep(0.5)
        pyautogui.hotkey('ctrl','s')
        # close_form()   
        time.sleep(1)
        pyautogui.hotkey('ctrl','f4')
        time.sleep(1)
        pyautogui.hotkey('ctrl','f4')
        time.sleep(1)
        # pyautogui.keyDown('Enter')
        return 'Y'
    
    except:
        messagebox.showerror("Error", "Inventory Asset update, 오류가 발생했습니다.")
        return 'N'

def pending_update():
    try:
        running = True
        label_result.config(text='Pending Cost Update...')
        wb = load_workbook(r'.\data\pending.xlsx', data_only= True, read_only=True)
        ws = wb['Sheet1']
        total_records = ws.max_row
        wait_time = float(pyautogui.prompt('대기 초는 얼마나 할까요? 기본은 0.5 이상을 입력 하세요.'))# type: ignore  
        if wait_time<= 0.4: 
            pyautogui.alert('0.4 이하를 입력 하였습니다. 프로그램을 다시 실행 하여 주세요. ')# type: ignore  
            exit()
        pyautogui.PAUSE = wait_time
        
        
        job = pyautogui.confirm(text=' Pending Cost Update 작업을 진행 하시겠습니끼?', buttons =['OK','Cancel'])# type: ignore  
        if job == 'OK':
            rec = 0
            time.sleep(3)
            global cost_element1
            global cost_element2
            global sub_element2 
            for row in ws.iter_rows():
                rec += 1
                
                item = str(row[0].value)
                cost_type = str(row[1].value)
                cost = str(row[2].value)
                overhead = str(row[3].value)
                
                update_progress(f'Pending Cost Update 진행 상황: {rec}/{total_records} , {item}')  # 현재 진행 상황 업데이트
                
                cost_menu_image_check(item_mp, item, cost_type)
                check_not_cost = item_cost_type_not_found()  
                
                if check_not_cost == 'Y':
                    #해당 cost type 이 존재 하지 않는 경우
                    check_new = item_new_cost(item, cost_type)
                    if check_new =='Y':
                    # new 이면서 inventory asset 인경우
                        cost_type_inventory(cost_type_inventory_img,cost, overhead )
                else:
                    #
                    ri= pyautogui.locateOnScreen(r'.\images\basedroll1.png')
                    time.sleep(0.5)
                    if ri is not None:
                        # print(item,' type : default, inventory, baseed rollup')
                        ric = pyautogui.center(ri) 
                        # defaul value uncheckㅡ
                        # time.sleep(0.3)
                        pyautogui.click(ric.x-130, ric.y)
                        # cost roll unckeck
                        # time.sleep(0.3)
                        pyautogui.click(ric.x+130, ric.y)
                        # time.sleep(1)
                        # cost zero 로 update
                        pyautogui.hotkey('enter')                              
                        # time.sleep(0.5)
                        pyautogui.hotkey('ctrl','s')
                        # time.sleep(1)
                        pyautogui.hotkey('tab')
                    else:
                        # inventory asset, cost rollup 인경우
                        time.sleep(1)
                        ri_undefault = pyautogui.locateOnScreen(r'.\images\item_cost_not_default.png')
                        if ri_undefault is not None:
                            # print(item,' type : not default, inventory asset , based on rollup')
                            ri_undefault_c = pyautogui.center(ri_undefault)                             
                            pyautogui.click(ri_undefault_c.x+115, ri_undefault_c.y+10)
                            pyautogui.hotkey('enter')
                            pyautogui.hotkey('ctrl','s')
                            pyautogui.hotkey('tab')     
                        
                    # check_inv_YN = cost_type_inventory(cost_type_inventory_img,cost, overhead )
                    check_cost_YN = cost_update_function(cost, overhead)
                    if check_cost_YN == 'Y':
                        update_progress(f'Pending Update 중  : {rec}/{total_records} , {item}') 
                        running = True
                    else:    
                        update_progress(f'Pending Update 중 Error : {rec}/{total_records} , {item}') 
                        running = False
                        break    
            if running == True:
                # cost update program 실행
                Frozen_update()
           
                    
                    
    except:
        messagebox.showerror('Error',' Pending Cost Update error')

# Frozen cost update 실행하기
def Frozen_update(value=None):
    try:
        wb = load_workbook(r'.\data\pending.xlsx', data_only= True, read_only=True)
        ws = wb['Sheet1']
        total_records = ws.max_row
        time.sleep(1)
        
        label_result.config(text='Item Cost Update...')
        if value == 'M':
             job = pyautogui.confirm(text=' Item Cost Update 작업을 진행 하시겠습니끼?', buttons =['OK','Cancel'])# type: ignore  
        else:
            job='OK'
        if job != 'OK':
            return 0
        
        rec = 0        
        std_menu = pyautogui.locateOnScreen(r'.\images\standard.png')
        # time.sleep(2)
      
        if std_menu != None:
            pyautogui.doubleClick(pyautogui.center(std_menu))     
            
            time.sleep(0.5)
                    
            for row_f in ws.iter_rows():
                rec += 1                
                 
                item_code = str(row_f[0].value)
                cost_type = str(row_f[1].value)
               
                # cost_type = 'Pending'
                # if cost_type not in ['Pending', 'Annual'] : break
                
                x = dt.datetime.now()                
                # print(x.strftime('%Y%m%d%H%M'))
                cost_remark = 'KRK_'+ item_code+'_'+x.strftime('%Y%m%d%H%M')            
                time.sleep(1)    
                # update_menu = pyautogui.locateOnScreen(resource_path(r'images\nov_cost_update.png'), confidence=0.8)
                # update_menu = pyautogui.locateOnScreen(r'.\images\nov_cost_update.png',confidence=0.8)
                update_menu = pyautogui.locateOnScreen(r'.\images\nov_cost_update.png', confidence=0.8)
                # time.sleep(2)
                if update_menu is not None:
                   
                    update_menu_cen = pyautogui.center(update_menu)
                    # time.sleep(0.5)
                    pyautogui.doubleClick(update_menu_cen)
                    
                    update_progress(f'Pending/Frozen Update 중  : {rec}/{total_records} {item_code}') 
                    # time.sleep(0.5)
                    pyautogui.press('tab', presses=2, interval=0.3)
                    time.sleep(0.3)
                    pyautogui.write(cost_type, interval=0.2)
                    pyautogui.hotkey('tab')
                    time.sleep(0.3)
                    pyautogui.write(cost_remark, interval=0.1)
                    time.sleep(0.3)
                    pyautogui.press('tab', presses=2, interval=0.3)
                    time.sleep(0.3)
                    pyautogui.write(item_code, interval=0.1)
                    time.sleep(0.3)
                    pyautogui.hotkey('tab')
                    time.sleep(0.3)
                    pyautogui.hotkey('enter')
                    time.sleep(0.3)
                    pyautogui.press('tab', presses=3, interval=0.3)
                    pyautogui.hotkey('enter')
                    time.sleep(2)                    
                    # summit_btn = pyautogui.locateOnScreen(resource_path(r'images\summit_btn.png'), confidence=0.8)
                    summit_btn = pyautogui.locateOnScreen(r'.\images\summit_btn.png')
                    # time.sleep(0.5)
                    if summit_btn is not None:
                        # print('btn click')
                        pyautogui.click(pyautogui.center(summit_btn)) 
                        time.sleep(1)
                        pyautogui.hotkey('tab')
                        time.sleep(0.5)
                        pyautogui.hotkey('enter')
                       
                else:
                    messagebox.showerror('Error',' NOV Update Costs Menu 못 찾음')   
                    # pyautogui.screenshot(cost_remark)
                    break
                
            pyautogui.alert('작업을 종료 하였습니다!!!!')     # type: ignore    
        
        
    except:
        messagebox.showerror('Error',' Frozen Cost Update error')          
                    
            
mainFrame = tk.Tk()
mainFrame.title('Item Cost 관련 Program')
width = 800  # 창의 너비
height = 400  # 창의 높이
x = (mainFrame.winfo_screenwidth() // 2) - (width // 2)  # 창의 가로 위치 (중앙 정렬)
y = (mainFrame.winfo_screenheight() // 2) - (height // 2)  # 창의 세로 위치 (중앙 정렬)

mainFrame.geometry(f"{width}x{height}+{x}+{y}")

item_mp =r'.\images\item_cost1.png'   
cost_3_yes = r'.\images\basedroll1.png' 
cost_not_default_img = r'.\images\item_cost_not_default.png' 
cost_type_inventory_img = r'.\images\inv_asset.png' 
cost_type_not_found = r'.\image\cost_type_not_found.png'
baseroll_img = r'.\image\baseroll.png'
close_img = r'\image\close.png'
submit_img = r'\image\summit_btn.png'

nov_update_costs_img = r'\image\nov_cost_update.png'
close1_img = r'\image\close1.png'
costs_btn_img = r'\image\costs_btn.png'
find_img = r'\image\find.png'
session_check_img = r'\image\session_check.png'
standard_img = r'\image\standard.png'
# image1 = ImageTk.PhotoImage(Image.open(r'D:\python_dev\item_cost_update\images\basedroll.png'))
# image2 = ImageTk.PhotoImage(Image.open(r'D:\python_dev\item_cost_update\images\basedroll1.png'))
# image3 = ImageTk.PhotoImage(Image.open(r'D:\python_dev\item_cost_update\images\close.png'))
# image4 = ImageTk.PhotoImage(Image.open(r'D:\python_dev\item_cost_update\images\item_cost1.png'))
# image5 = ImageTk.PhotoImage(Image.open(r'D:\python_dev\item_cost_update\images\screen_max.png'))

cost_element1 = 'Material'
cost_element2 = 'Material Overhead'
sub_element2 = 'Freight'


lbl1 = tk.Label(mainFrame, text=' Standard Cost 관련 작업입니다. ', fg='red')# type: ignore
lbl1.grid(row=0, column=0)
lbl2 = tk.Label(mainFrame, text=' 해당 프로그램을 ERP 화면과 중복되지 않도록 이동하세요 !!!', fg='red')# type: ignore
lbl2.grid(row=1, column=0)


btn1 = tk.Button(mainFrame, text='Annual Cost Reset',width=30,height=2,  command=annual_reset, activebackground='green')# type: ignore
btn2 = tk.Button(mainFrame, text='Annual Cost Update',width=30,height=2,  command=annual_update, activebackground='green')# type: ignore
btn3 = tk.Button(mainFrame, text='Pending /Frozen Cost Update',width=30,height=2,command=pending_update, activebackground='green' )# type: ignore
btn4 = tk.Button(mainFrame, text='Frozen Cost Update',width=30,height=2 ,command=lambda: Frozen_update('M'), activebackground='green')# type: ignore

btn1.grid(row=2, column=0)
btn2.grid(row=3, column=0)
btn3.grid(row=4, column=0)
btn4.grid(row=5, column=0)

mainFrame.bind('<Escape>', on_escape)

progress_label = tk.Label(mainFrame, text='진행 상황', fg='blue')
progress_label.grid(row=2, column=1)

label_result = tk.Label(mainFrame, text='', fg='green')
label_result.grid(row=3, column=1)
lbl3 = tk.Label(mainFrame, text='Annual 초기화 관련 파일명 : ANNUAL_list.xlsx', fg='blue')
lbl4 = tk.Label(mainFrame, text='Annual Update 관련 파일명 : ANNUAL_UPDATE.xlsx', fg='blue') 
lbl5 = tk.Label(mainFrame, text= 'Pending Cost 관련은 : pending.xlsx' ,fg='blue')

lbl3.grid(row=6, column=0)
lbl4.grid(row=7, column=0)
lbl5.grid(row=8, column=0)

mainFrame.mainloop()        