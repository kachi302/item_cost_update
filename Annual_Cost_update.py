from cgitb import text
import sys
import os
import pyautogui
import time
import datetime as dt
from openpyxl import load_workbook


#print('Annual Item Cost Update Process, 5초후 시작 됩니다. 해당 화면을 다른 모니터로 이동 하세요.')

time.sleep(5)

# def resource_path(relative_path):
#     try:
#         base_path = sys._MEIPASS        
#     except Exception: 
#         base_path = os.path.abspath(".")

#     return os.path.join(base_path, relative_path)
                        
running = True    
rec=0
try:
    print('Start - 화면 크기 max 로')
    pyautogui.click(1000, 800)
    time.sleep(0.5)
    # screen_max = pyautogui.locateOnScreen(resource_path(r'images\screen_max.png'), confidence=0.8)
    screen_max = pyautogui.locateOnScreen(r'.\images\screen_max.png',confidence=0.9)
    
    
    if screen_max != None:
        screen_max_ce = pyautogui.center(screen_max)
        # pyautogui.moveTo(screen_max_ce.x , screen_max_ce.y-20)
        pyautogui.click(screen_max_ce.x , screen_max_ce.y-20, interval=0.5)
        print('Max Screen')
    else:
        print('Screen Max not found!~')        
     
    print('Excel File 읽기')           
    wb = load_workbook(r'.\data\pending.xlsx', data_only= True)
    ws = wb['Sheet1']
    total_records = ws.max_row
    # 전체 작업 시간
    start1 = time.time()
    job = pyautogui.confirm(text=' Annual Cost Update 작업을 진행 하시겠습니끼?', buttons =['OK','Cancel'])# type: ignore
    
    if job == 'OK':
        rec = 0
        time.sleep(3)
        cost_element1 = 'Material'
        cost_element2 = 'Material Overhead' 
        sub_element2 = 'Freight'
        sub_element3 ='Total Value'
        for row in ws.iter_rows():
            rec += 1
            # item별 작업 시간
            start = time.time()
            item = str(row[0].value)
            cost_type = str(row[1].value)   
            cost = str(row[2].value)
            overhead = str(row[3].value)
            print(item, cost_type, cost, overhead, str(total_records),str(rec))
            time.sleep(2)
            # item_mp = pyautogui.locateOnScreen(resource_path(r'images\item_cost1.png'))    
            item_mp = pyautogui.locateOnScreen(r'.\images\item_cost1.png', confidence=0.6)
            time.sleep(2)
            if item_mp != None:
                center=pyautogui.center(item_mp)
                pyautogui.doubleClick(center)
                time.sleep(1)                
                pyautogui.typewrite(item, interval=0.1)
                time.sleep(1)
                pyautogui.hotkey('tab')
                time.sleep(0.5)
                # cost type 입력
                pyautogui.typewrite(cost_type, interval=0.1)  
                time.sleep(0.5)
                pyautogui.hotkey('tab')  
                # Find Button Click
                time.sleep(0.5)
                # fi = pyautogui.locateOnScreen(resource_path(r'images\find.png'))
                fi = pyautogui.locateOnScreen(r'.\images\find.png')
                fic = pyautogui.center(fi)# type: ignore
                pyautogui.click(fic)                
                time.sleep(2)      
                
                # find click 했는데 Pending 이 없는 경우 
                # type_not_found = pyautogui.locateOnScreen(resource_path(r'images\cost_type_not_found.png'))
                type_not_found = pyautogui.locateOnScreen(r'.\images\cost_type_not_found_1.png', confidence=0.7)
                time.sleep(2)
                if type_not_found != None:
                    # pending cost type 이 없는 경우/ New Button click
                    print(item, '해당 cost type 이 없는 경우')
                    time.sleep(1)
                    # new_pending = pyautogui.locateOnScreen(resource_path(r'images\new.png'))
                    new_pending = pyautogui.locateOnScreen(r'.\images\new.png')
                    time.sleep(0.5)
                    if new_pending != None:
                        time.sleep(0.5)
                        new_center = pyautogui.center(new_pending)
                        pyautogui.click(new_center)
                        time.sleep(1)
                        # item code d
                        pyautogui.typewrite(item, interval=0.1)
                        time.sleep(1)
                        pyautogui.hotkey('tab')
                        time.sleep(0.5)
                        pyautogui.typewrite(cost_type, interval=0.1)  
                        time.sleep(0.5)
                                          
                        # inventory asset 인 경우
                        # inv_asset = pyautogui.locateOnScreen(resource_path(r'images\inv_asset.png'), grayscale=True, confidence=0.8)
                        inv_asset = pyautogui.locateOnScreen(r'.\images\inv_asset.png', grayscale=True, confidence=0.8)
                        time.sleep(1)
                        if  inv_asset != None :
                            # print('inventory asset')
                            time.sleep(0.5)
                            # cost btn click
                            # cost_btn = pyautogui.locateOnScreen(resource_path(r'images\costs_btn.png'))
                            cost_btn = pyautogui.locateOnScreen(r'.\images\costs_btn.png')
                            cost_btn_center=pyautogui.center(cost_btn)# type: ignore
                            pyautogui.click(cost_btn_center)
                            time.sleep(0.5)
                        #     pyautogui.typewrite(cost_element1, interval=0.1)
                            
                        #     time.sleep(0.5)
                        #     pyautogui.press('tab')
                        #     time.sleep(0.5)
                        #     pyautogui.typewrite(cost_element1, interval=0.1)
                        #     time.sleep(0.8)  
                        #     pyautogui.press('tab')
                        #     time.sleep(0.8)
                        #     pyautogui.press('tab')
                        #     time.sleep(0.8)                
                        #     #pyautogui.press('tab', presses=2, interval=1) 
                        #     #time.sleep(0.5)
                        #     # Material cost 입력
                        #     pyautogui.typewrite(str(cost), interval=0.1)
                        #     time.sleep(0.5)
                        #     pyautogui.hotkey('tab')
                        #     time.sleep(0.5)
                        #     pyautogui.hotkey('tab')
                        #     time.sleep(0.8)
                        #     # material overhead
                        #     pyautogui.typewrite(cost_element2, interval=0.1)
                        #     time.sleep(0.5)
                        #     pyautogui.hotkey('tab')
                        #     time.sleep(0.8)
                        #     # sub element
                        #     pyautogui.typewrite(sub_element2, interval=0.1)
                        #     time.sleep(0.8)
                        #     pyautogui.hotkey('tab')
                        #     time.sleep(0.8)
                        #     pyautogui.press('tab') 
                        #     time.sleep(1)
                        #     pyautogui.typewrite(sub_element3, interval=0.2)
                        #     time.sleep(0.5)
                        # # Material cost 입력         
                        #     pyautogui.hotkey('tab')
                        #     time.sleep(0.5)
                        #     pyautogui.typewrite(str(overhead), interval=0.1)
                        #     time.sleep(0.5)
                        #     pyautogui.hotkey('tab')                      
                            # 저장
                        #keybord 선택 하는 부분으로 변경
                            #cost element
                            pyautogui.hotkey('ctrl','l')
                            time.sleep(0.5)
                            pyautogui.hotkey('enter')
                            time.sleep(0.5)
                            #sub element, Material
                            pyautogui.hotkey('ctrl','l')
                            time.sleep(0.5)
                            #pyautogui.hotkey('enter')
                            #위 선택해면 자동으로 basis 로 넘어간다.
                            time.sleep(0.5)
                            pyautogui.hotkey('tab')
                            time.sleep(0.5)
                            #basis 부분에서 rate 부분으로 이동
                            pyautogui.hotkey('tab')
                            time.sleep(0.5)
                            # Material cost 입력
                            pyautogui.typewrite(str(cost), interval=0.1)
                            time.sleep(0.5)
                            pyautogui.hotkey('tab')
                            time.sleep(0.5)
                            pyautogui.hotkey('tab')
                            time.sleep(0.5)
                            # material overhead
                            pyautogui.hotkey('ctrl','l')
                            time.sleep(0.7)
                            pyautogui.press('down')        
                            time.sleep(0.5)  
                            pyautogui.hotkey('enter')        
                            time.sleep(0.5)  
                            #sub element(Freight)
                            #pyautogui.hotkey('tab')
                            #time.sleep(0.5)
                            pyautogui.hotkey('ctrl','l')
                            time.sleep(0.7)
                            pyautogui.press('down')        
                            time.sleep(0.5)  
                            pyautogui.hotkey('enter')        
                            time.sleep(1)  
                            # activity
                            pyautogui.hotkey('tab')
                            #basis
                            time.sleep(0.8)
                            pyautogui.hotkey('ctrl','l')
                            time.sleep(0.5)
                            pyautogui.press('down',presses=5)        
                            time.sleep(0.5)  
                            pyautogui.hotkey('enter')        
                            time.sleep(1)  
                            # rate
                            pyautogui.typewrite(str(overhead), interval=0.1)
                            
                            #pyautogui.hotkey('tab')    
                        
                        #----------------
                            #time.sleep(0.5)
                            #pyautogui.hotkey('tab')    
                            
                            time.sleep(0.5)
                            pyautogui.hotkey('ctrl','s')
                            time.sleep(0.5)
                            pyautogui.hotkey('ctrl','f4')
                            time.sleep(0.5)
                            pyautogui.hotkey('ctrl','f4')
                            time.sleep(1)
                            
                            print(item ,'Invetnroy Asset')
                            ws['E'+str(rec)]='Inventory Asset Completed'
                            print(item," 코드 실행 시간 :", time.time() - start)  
                            time.sleep(2)
                        else:
                            print('Inventory Asset 이 아닌 경우')  
                            ws['E'+str(rec)]='Inventory Non Asset item'  
            
                else:
                    # pending cost type 이 있는 경우                 
                    # inventory asset 인 경우
                    print(item, '해당 cost type 이 있는 경우')
                    time.sleep(1)
                    # inv_asset = pyautogui.locateOnScreen(resource_path(r'images\inv_asset.png'))
                    inv_asset = pyautogui.locateOnScreen(r'.\images\inv_asset.png')
                    time.sleep(2)
                    if  inv_asset != None :
                        # print('inventory asset')
                        print(item ,'Annual cost 있는 경우 Invetnroy Asset')
                        time.sleep(0.5)
                        # cost btn click
                        # cost_btn = pyautogui.locateOnScreen(resource_path(r'images\costs_btn.png'))
                        cost_btn = pyautogui.locateOnScreen(r'.\images\costs_btn.png')
                        cost_btn_center=pyautogui.center(cost_btn)# type: ignore
                        pyautogui.click(cost_btn_center)
                        time.sleep(2)
                        # pyautogui.typewrite(cost_element1, interval=0.1)
                        # time.sleep(0.8)
                        # #pyautogui.hotkey('tab')
                        # pyautogui.press('tab')
                        # time.sleep(0.8)
                        # pyautogui.typewrite(cost_element1, interval=0.1)
                        # time.sleep(0.8)                        
                        # #pyautogui.press('tab', presses=2, interval=0.5)
                        # #pyautogui.hotkey('tab')
                        # pyautogui.press('tab')
                        # time.sleep(0.8)
                        # #pyautogui.hotkey('tab')
                        # pyautogui.press('tab')
                        # time.sleep(0.8)                        
                        # pyautogui.typewrite('Item', interval=0.1) 
                        # time.sleep(0.8)
                        # #pyautogui.hotkey('tab')
                        # pyautogui.press('tab')
                        # time.sleep(0.8)
                        # # Material cost 입력
                        # pyautogui.typewrite(str(cost), interval=0.1)
                        # time.sleep(0.8)
                        # #pyautogui.hotkey('tab')
                        # pyautogui.press('tab')
                        # time.sleep(0.8)
                        # #pyautogui.hotkey('tab')
                        # pyautogui.press('tab')
                        # time.sleep(0.8)
                        # # material overhead
                        # pyautogui.typewrite(cost_element2, interval=0.1)
                        # time.sleep(0.8)
                        # #pyautogui.hotkey('tab')
                        # pyautogui.press('tab')
                        # time.sleep(0.8)
                        # # sub element
                        # pyautogui.typewrite(sub_element2, interval=0.1)
                        # time.sleep(0.8)
                        # #pyautogui.hotkey('tab')
                        # pyautogui.press('tab')
                        # time.sleep(0.8)
                        # pyautogui.press('tab') 
                        # time.sleep(1)
                        # pyautogui.typewrite(sub_element3, interval=0.2)
                        # time.sleep(0.8)
                        # # Material cost 입력         
                        # #pyautogui.hotkey('tab')
                        # pyautogui.press('tab')
                        # time.sleep(0.8)
                        # pyautogui.typewrite(str(overhead), interval=0.1)
                        # time.sleep(0.5)
                        # #pyautogui.hotkey('tab')                        
                        # pyautogui.press('tab')
                        # 저장
                        #----------------Material
                        pyautogui.hotkey('ctrl','l')
                        time.sleep(0.5)
                        pyautogui.hotkey('enter')
                        time.sleep(0.5)
                        #sub element, Material
                        pyautogui.hotkey('ctrl','l')
                        time.sleep(0.5)
                        #pyautogui.hotkey('enter')
                        #위 선택해면 자동으로 basis 로 넘어간다.
                        time.sleep(0.5)
                        pyautogui.hotkey('tab')
                        time.sleep(0.5)
                        #basis 부분에서 rate 부분으로 이동
                        pyautogui.hotkey('tab')
                        time.sleep(0.5)
                        # Material cost 입력
                        pyautogui.typewrite(str(cost), interval=0.1)
                        time.sleep(0.5)
                        pyautogui.hotkey('tab')
                        time.sleep(0.7)
                        pyautogui.hotkey('tab')
                        time.sleep(0.8)
                        # material overhead
                        pyautogui.hotkey('ctrl','l')
                        time.sleep(0.5)
                        pyautogui.press('down')        
                        time.sleep(0.5)  
                        pyautogui.hotkey('enter')        
                        time.sleep(0.5)  
                        #sub element(Freight)
                        #pyautogui.hotkey('tab')
                        #time.sleep(0.5)
                        pyautogui.hotkey('ctrl','l')
                        time.sleep(0.7)
                        pyautogui.press('down')        
                        time.sleep(0.5)  
                        pyautogui.hotkey('enter')        
                        time.sleep(0.7)  
                        # activity
                        pyautogui.hotkey('tab')
                        #basis
                        time.sleep(0.5)
                        pyautogui.hotkey('ctrl','l')
                        time.sleep(0.9)
                        pyautogui.press('down',presses=5)        
                        time.sleep(0.5)  
                        pyautogui.hotkey('enter')        
                        time.sleep(0.5)  
                        # rate
                        pyautogui.typewrite(str(overhead), interval=0.1)
                        time.sleep(0.5)
                        #pyautogui.hotkey('tab')    
                        
                        #----------------
                        
                        #time.sleep(0.5)
                        pyautogui.hotkey('ctrl','s')    
                        time.sleep(0.5)
                        
                        ws['E'+str(rec)]='Inventory Asset Completed'
                        print(item," Annual cost update 실행 시간 :", time.time() - start)  
                                                
                        pyautogui.hotkey('ctrl','f4')
                        time.sleep(0.5)
                        pyautogui.hotkey('ctrl','f4')
                        time.sleep(0.5)
                        
                        
                    else:
                        # Annual cost type 이 있는 경우
                        # default control 인 경우, 값 초기화 하기
                        time.sleep(0.5)
                        # ri= pyautogui.locateOnScreen(resource_path(r'images\basedroll1.png'), confidence=0.8)
                        ri= pyautogui.locateOnScreen(r'.\images\basedroll1.png')
                        time.sleep(1)
                        if ri != None:
                            print(item,' type : 3가지 다 있는 경우,default ')
                            ric = pyautogui.center(ri) 
                            # defaul value uncheck
                            time.sleep(1)
                            pyautogui.click(ric.x-130, ric.y)
                            # cost roll unckeck
                            time.sleep(0.5)
                            pyautogui.click(ric.x+130, ric.y)
                            time.sleep(1)
                            # cost zero 로 update
                            pyautogui.hotkey('enter')                              
                            time.sleep(1)
                            pyautogui.hotkey('ctrl','s')
                            time.sleep(1)
                            pyautogui.hotkey('tab')
                            ws['B'+str(rec)]='type : 3가지 다 있는 경우,default ,CHANGE Completed'    
                        else:
                                
                            time.sleep(1)
                            # ri_undefault = pyautogui.locateOnScreen(resource_path(r'images\item_cost_not_default.png'))
                            ri_undefault = pyautogui.locateOnScreen(r'.\images\item_cost_not_default.png')
                            time.sleep(1)
                            if ri_undefault != None:
                                print(item,' type : Inventory, Based on Rollup')
                                ri_undefault_c = pyautogui.center(ri_undefault) 
                                time.sleep(1)  
                                pyautogui.click(ri_undefault_c.x+115, ri_undefault_c.y+10)
                                time.sleep(1)
                                pyautogui.hotkey('enter')
                                time.sleep(0.5)                    
                                pyautogui.hotkey('ctrl','s')
                                # close_form()   
                                time.sleep(1)
                                pyautogui.hotkey('tab')            
                                
                                ws['B'+str(rec)]='type : Inventory, Based on Rollup, Completed'    
                        # cost update, cost type 을 inventory asset 으로 변경 후 cost update
                        time.sleep(0.5)
                        # cost btn click
                        # cost_btn = pyautogui.locateOnScreen(resource_path(r'images\costs_btn.png'))
                        cost_btn = pyautogui.locateOnScreen(r'.\images\costs_btn.png')
                        cost_btn_center=pyautogui.center(cost_btn)# type: ignore
                        pyautogui.click(cost_btn_center)
                        time.sleep(1)
                        # pyautogui.typewrite(cost_element1, interval=0.1)
                        # time.sleep(0.5)
                        # #pyautogui.hotkey('tab')
                        # pyautogui.press('tab')
                        # time.sleep(0.8)
                        # pyautogui.typewrite(cost_element1, interval=0.1)
                        # time.sleep(0.8)                        
                        # ##pyautogui.press('tab', presses=2, interval=0.3)
                        # #pyautogui.hotkey('tab')
                        # pyautogui.press('tab')
                        # time.sleep(0.8)
                        
                        # #pyautogui.hotkey('tab')
                        # pyautogui.press('tab')
                        # time.sleep(0.8)
                        
                        # pyautogui.typewrite('Item', interval=0.1) 
                        # time.sleep(0.8)
                        # #pyautogui.hotkey('tab')
                        # pyautogui.press('tab')
                        # time.sleep(0.8)
                        # # Material cost 입력
                        # pyautogui.typewrite(str(cost), interval=0.1)
                        # time.sleep(0.5)
                        # #pyautogui.hotkey('tab')
                        # pyautogui.press('tab')
                        # time.sleep(0.8)
                        # #pyautogui.hotkey('tab')
                        # pyautogui.press('tab')
                        # # material overhead
                        # time.sleep(0.8)
                        # pyautogui.typewrite(cost_element2, interval=0.1)
                        # time.sleep(0.5)
                        # #pyautogui.hotkey('tab')
                        # pyautogui.press('tab')
                        # time.sleep(0.8)
                        # # sub element
                        # pyautogui.typewrite(sub_element2, interval=0.1)
                        # time.sleep(0.8)
                        # #pyautogui.hotkey('tab')
                        # pyautogui.press('tab')
                        # time.sleep(0.8)
                        # pyautogui.press('tab') 
                        # time.sleep(1)
                        # pyautogui.typewrite(sub_element3, interval=0.2)
                        # time.sleep(0.8)
                        # # Material cost 입력         
                        # #pyautogui.hotkey('tab')
                        # pyautogui.press('tab')
                        # time.sleep(0.5)
                        # pyautogui.typewrite(str(overhead), interval=0.1)
                        # time.sleep(0.5)
                        # #pyautogui.hotkey('tab')                        
                        # pyautogui.press('tab')
                        #----------------
                        pyautogui.hotkey('ctrl','l')
                        time.sleep(0.5)
                        pyautogui.hotkey('enter')
                        time.sleep(0.5)
                        #sub element, Material
                        pyautogui.hotkey('ctrl','l')
                        time.sleep(0.5)
                        #pyautogui.hotkey('enter')
                        #위 선택해면 자동으로 basis 로 넘어간다.
                        time.sleep(0.5)
                        pyautogui.hotkey('tab')
                        time.sleep(0.5)
                        #basis 부분에서 rate 부분으로 이동
                        pyautogui.hotkey('tab')
                        time.sleep(0.5)
                        # Material cost 입력
                        pyautogui.typewrite(str(cost), interval=0.1)
                        time.sleep(0.5)
                        pyautogui.hotkey('tab')
                        time.sleep(0.7)
                        pyautogui.hotkey('tab')
                        time.sleep(0.8)
                        # material overhead
                        pyautogui.hotkey('ctrl','l')
                        time.sleep(0.5)
                        pyautogui.press('down')        
                        time.sleep(0.5)  
                        pyautogui.hotkey('enter')        
                        time.sleep(0.5)  
                        #sub element(Freight)
                        #pyautogui.hotkey('tab')
                        time.sleep(0.5)
                        pyautogui.hotkey('ctrl','l')
                        time.sleep(0.5)
                        pyautogui.press('down')        
                        time.sleep(0.5)  
                        pyautogui.hotkey('enter')        
                        time.sleep(0.7)  
                        # activity
                        pyautogui.hotkey('tab')
                        #basis
                        time.sleep(0.7)
                        pyautogui.hotkey('ctrl','l')
                        time.sleep(0.5)
                        pyautogui.press('down',presses=5)        
                        time.sleep(0.5)  
                        pyautogui.hotkey('enter')        
                        time.sleep(0.5)  
                        # rate
                        pyautogui.typewrite(str(overhead), interval=0.1)
                        #time.sleep(0.5)
                        #pyautogui.hotkey('tab')    
                       
                        #----------------
                        # 저장
                        time.sleep(0.8)
                        pyautogui.hotkey('ctrl','s')
                        time.sleep(0.58)
                        pyautogui.hotkey('ctrl','f4')
                        time.sleep(0.55)
                        pyautogui.hotkey('ctrl','f4')
                        time.sleep(1)
                        
                        #print(item ,'Annual cost 있는 경우 Invetnroy Asset')
                        ws['E'+str(rec)]=' Inventory Asset Completed'
                        print(item," Annual Cost Update 실행 시간 :", time.time() - start)      
            else:    
                print(item, cost_type, cost, overhead, str(total_records),str(rec))
                break
       
                    
    print(" END 코드 실행 시간 :", time.time() - start1, rec)   
    wb.active = ws
    # wb.save(resource_path(r'data\Pending_JOB.xlsx')) 
    wb.save(r'.\data\Annual_LIST_JOB.xlsx')       
    pyautogui.alert('작업을 종료 하였습니다!!!!')        # type: ignore
except Exception as e:
    # ws['E'+str(rec)]=e
    # wb.active = ws    
    # wb.save(resource_path(r'data\Pending_JOB.xlsx'))   
    wb.active = ws
    # wb.save(resource_path(r'data\Pending_JOB.xlsx')) 
    wb.save(r'.\data\Annual_LIST_JOB.xlsx')   
    
    print(e)    
    pyautogui.alert(text=e, button='OK')# type: ignore
    time.sleep(10)    
   