import sys
import os
import pyautogui
import time
import datetime as dt
from openpyxl import load_workbook
pyautogui.FAILSAFE = False


print('Item Cost Update Process, 5초후 시작 됩니다. 해당 화면을 다른 모니터로 이동 하세요.')
time.sleep(5)

# def resource_path(relative_path):
#     try:
#         base_path = sys._MEIPASS        
#     except Exception: 
#         base_path = os.path.abspath(".")

#     return os.path.join(base_path, relative_path)
                        
running = True    
rec=0
try:
    print('Start - 화면 크기 max 로')
    pyautogui.click(1000, 800)
    # time.sleep(0.5)
    # ---screen_max = pyautogui.locateOnScreen(resource_path(r'images\screen_max.png'), confidence=0.8)
    screen_max = pyautogui.locateOnScreen(r'.\images\screen_max.png',confidence=0.9)
        
    if screen_max != None:
        screen_max_ce = pyautogui.center(screen_max)
        # pyautogui.moveTo(screen_max_ce.x , screen_max_ce.y-20)
        pyautogui.click(screen_max_ce.x , screen_max_ce.y-20, interval=1)
        print('Max Screen')
    else:
        print('Screen Max not found!~')        
     
    print('Excel File 읽기')           
    wb = load_workbook(r'.\data\pending.xlsx', data_only= True)
    ws = wb['Sheet1']
    # w_cost = wb['Sheet1']
    # 전체 작업 시간
    start1 = time.time()
    wait_time = float(pyautogui.prompt('대기 초는 얼마나 할까요? 기본은 0.5 이상을 입력 하세요.'))# type: ignore  
    if wait_time<= 0.4: 
        pyautogui.alert('0.4 이하를 입력 하였습니다. 프로그램을 다시 실행 하여 주세요. ')# type: ignore  
        exit()
    pyautogui.PAUSE = wait_time
    print('대기 시간은 ', wait_time)
    job = pyautogui.confirm(text=' Pending Cost Update 작업을 진행 하시겠습니끼?', buttons =['OK','Cancel'])# type: ignore  
    
    
    if job == 'OK':
        rec = 0
        time.sleep(3)
        cost_element1 = 'Material'
        cost_element2 = 'Material Overhead'
        sub_element2 = 'Freight'
        
        for row in ws.iter_rows():
            rec += 1
            # item별 작업 시간
            start = time.time()
            item = str(row[0].value)
            cost_type = str(row[1].value)
            cost = str(row[2].value)
            overhead = str(row[3].value)
            # print(item, cost_type, cost, overhead)
            time.sleep(1)
            # item_mp = pyautogui.locateOnScreen(resource_path(r'images\item_cost1.png'))    
            
            item_mp = pyautogui.locateOnScreen(r'.\images\item_cost1.png')
            if item_mp is not None:
                center=pyautogui.center(item_mp)
                pyautogui.doubleClick(center)
                # time.sleep(0.5)
                
                pyautogui.write(item, interval=0.1)
                # time.sleep(0.7)
                pyautogui.hotkey('tab')
                # time.sleep(0.7)
                # cost type 입력
                pyautogui.write(cost_type, interval=0.1)  
                # time.sleep(0.7)
                pyautogui.hotkey('tab')  
                # Find Button Click
                # time.sleep(0.7)
                # fi = pyautogui.locateOnScreen(resource_path(r'images\find.png'))
                fi = pyautogui.locateOnScreen(r'.\images\find.png')
                if fi is not None:
                    fic = pyautogui.center(fi)
                    pyautogui.click(fic)                
                    time.sleep(1)      
                    
                    # find click 했는데 Pending 이 없는 경우 
                    # type_not_found = pyautogui.locateOnScreen(resource_path(r'images\cost_type_not_found.png'))
                    type_not_found = pyautogui.locateOnScreen(r'.\images\cost_type_not_found.png')
                    if type_not_found is not None:
                        # pending cost type 이 없는 경우/ New Button click
                        time.sleep(1)
                        # new_pending = pyautogui.locateOnScreen(resource_path(r'images\new.png'))
                        new_pending = pyautogui.locateOnScreen(r'.\images\new.png')
                        time.sleep(0.5)
                        if new_pending is not None:
                            new_center = pyautogui.center(new_pending)
                            pyautogui.click(new_center)
                            # time.sleep(1)
                            # item code d
                            pyautogui.write(item, interval=0.1)
                            # time.sleep(0.5)
                            pyautogui.hotkey('tab')
                            pyautogui.write(cost_type, interval=0.1)  
                            # time.sleep(0.5)
                                            
                            # inventory asset 인 경우
                            # inv_asset = pyautogui.locateOnScreen(resource_path(r'images\inv_asset.png'), grayscale=True, confidence=0.8)
                            inv_asset = pyautogui.locateOnScreen(r'.\images\inv_asset.png')
                            if  inv_asset is not None:
                                # print('inventory asset')
                                # time.sleep(0.5)
                                # cost btn click
                                # cost_btn = pyautogui.locateOnScreen(resource_path(r'images\costs_btn.png'))
                                cost_btn = pyautogui.locateOnScreen(r'.\images\costs_btn.png')
                                # time.sleep(0.5)
                                if cost_btn is None: break
                                cost_btn_center=pyautogui.center(cost_btn)
                                # time.sleep(0.3)
                                pyautogui.click(cost_btn_center)
                                # time.sleep(0.5)
                                pyautogui.write(cost_element1, interval=0.1)
                                # time.sleep(0.3)
                                pyautogui.hotkey('tab')
                                # time.sleep(0.2)
                                pyautogui.write(cost_element1, interval=0.1)
                                # time.sleep(0.3)                        
                                pyautogui.press('tab', presses=3, interval=0.5) 
                                # time.sleep(0.5)
                                # Material cost 입력
                                pyautogui.write(cost, interval=0.1)
                                # time.sleep(0.5)
                                pyautogui.hotkey('tab')
                                # time.sleep(0.5)
                                pyautogui.hotkey('tab')
                                # time.sleep(0.2)
                                # material overhead
                                pyautogui.write(cost_element2, interval=0.1)
                                # time.sleep(0.2)
                                pyautogui.hotkey('tab')
                                # sub element
                                # time.sleep(0.5)
                                pyautogui.write(sub_element2, interval=0.1)
                                
                                # time.sleep(0.2)
                                pyautogui.hotkey('tab')
                                # time.sleep(0.3)
                                pyautogui.press('tab', presses=2, interval=0.3) 
                                # Material Overhead rate 입력
                                # time.sleep(0.5)                        
                                pyautogui.write(overhead, interval=0.1)
                                # time.sleep(0.2)
                                pyautogui.hotkey('tab')                        
                                # 저장
                                # time.sleep(0.3)
                                pyautogui.hotkey('ctrl','s')
                                # time.sleep(0.5)
                                pyautogui.hotkey('ctrl','f4')
                                # time.sleep(0.5)
                                pyautogui.hotkey('ctrl','f4')
                                # time.sleep(0.5)
                                
                                print(item ,'Invetnroy Asset')
                                ws['E'+str(rec)]='Inventory Asset Completed'
                                print(item," 코드 실행 시간 :", time.time() - start)  
                            else:
                                print('Inventory Asset 이 아닌 경우')
                                ws['E'+str(rec)]='Inventory Non Asset item'  
                
                    else:
                        # pending cost type 이 있는 경우                 
                        # inventory asset 인 경우
                        time.sleep(1)
                        # inv_asset = pyautogui.locateOnScreen(resource_path(r'images\inv_asset.png'))'
                        # inventory asset 만 check 되어 있는 경우
                        inv_asset = pyautogui.locateOnScreen(r'.\images\inv_asset.png')
                        if  inv_asset is not None :
                            # print('inventory asset')
                            time.sleep(0.5)
                            # cost btn click
                            # cost_btn = pyautogui.locateOnScreen(resource_path(r'images\costs_btn.png'))
                            cost_btn = pyautogui.locateOnScreen(r'.\images\costs_btn.png')
                            if cost_btn is None: break
                            cost_btn_center=pyautogui.center(cost_btn)
                            pyautogui.click(cost_btn_center)
                            # time.sleep(0.5)
                            # cost information 화면 뜨고
                            # element 항목 넣기(Material)
                            pyautogui.write(cost_element1, interval=0.1)
                            # time.sleep(0.2)
                            pyautogui.hotkey('tab')
                            #sub element 항목 넣기
                            pyautogui.write(cost_element1, interval=0.1)
                            # time.sleep(0.2)       
                            # activity 칼럼                 
                            pyautogui.press('tab')
                            # time.sleep(0.3)
                            # basis 칼럼으로 이동
                            pyautogui.press('tab')
                            # time.sleep(0.3)
                            pyautogui.write('Item', interval=0.1) 
                            # time.sleep(0.3)
                            pyautogui.hotkey('tab')
                            # time.sleep(0.2)
                            # Material cost 입력
                            pyautogui.write(cost, interval=0.1)
                            pyautogui.hotkey('tab')
                            # time.sleep(0.5)
                            pyautogui.hotkey('tab')
                            # material overhead
                            pyautogui.write(cost_element2, interval=0.1)
                            # time.sleep(0.2)
                            pyautogui.hotkey('tab')
                            # sub element
                            pyautogui.write(sub_element2, interval=0.1)
                            # time.sleep(0.4)
                            pyautogui.hotkey('tab')
                            # time.sleep(0.5)
                            pyautogui.press('tab') 
                            # time.sleep(0.5)
                            pyautogui.write('Total Value', interval=0.1)
                            # Material cost 입력         
                            pyautogui.hotkey('tab')
                            # time.sleep(0.3)
                            pyautogui.write(overhead, interval=0.1)
                            # time.sleep(0.2)
                            pyautogui.hotkey('tab')                        
                            # 저장
                            # time.sleep(0.3)
                            pyautogui.hotkey('ctrl','s')    
                            # time.sleep(0.5)
                            pyautogui.hotkey('ctrl','f4')
                            # time.sleep(0.5)
                            pyautogui.hotkey('ctrl','f4')
                            # time.sleep(0.5)
                            
                            print(item ,'peingind cost 있는 경우 Invetnroy Asset')
                            ws['E'+str(rec)]='Inventory Asset Completed'
                            print(item," Pending cost update 실행 시간 :", time.time() - start)  
                        else:
                            # Pending cost type 이 있는 경우
                            # default control 인 경우, 값 초기화 하기, 3가지 다 있는 경우 획인(user default controls / Inventory Asset / Baseed on Rollup)
                            time.sleep(0.5)
                            # ri= pyautogui.locateOnScreen(resource_path(r'images\basedroll1.png'), confidence=0.8)
                            
                            ri= pyautogui.locateOnScreen(r'.\images\basedroll1.png')
                            time.sleep(0.5)
                            if ri is not None:
                                print(item,' type : default, inventory, baseed rollup')
                                ric = pyautogui.center(ri) 
                                # defaul value uncheckㅡ
                                # time.sleep(0.3)
                                pyautogui.click(ric.x-130, ric.y)
                                # cost roll unckeck
                                # time.sleep(0.3)
                                pyautogui.click(ric.x+130, ric.y)
                                # time.sleep(1)
                                # cost zero 로 update
                                pyautogui.hotkey('enter')                              
                                # time.sleep(0.5)
                                pyautogui.hotkey('ctrl','s')
                                # time.sleep(1)
                                pyautogui.hotkey('tab')
                            else:
                                    
                                time.sleep(1)
                                # 2가지인 경우 (inventory asset / Baseed on Rollup)
                                # ri_undefault = pyautogui.locateOnScreen(resource_path(r'images\item_cost_not_default.png'))
                                ri_undefault = pyautogui.locateOnScreen(r'.\images\item_cost_not_default.png')
                                if ri_undefault is not None:
                                    print(item,' type : not default, inventory asset , based on rollup')
                                    ri_undefault_c = pyautogui.center(ri_undefault) 
                                    # time.sleep(0.5)  
                                    # pyautogui.click(ri_undefault_c.x+115, ri_undefault_c.y+10)
                                    #20230220 좌표 수정
                                    pyautogui.click(ri_undefault_c.x+115, ri_undefault_c.y+10)
                                    # time.sleep(1)
                                    pyautogui.hotkey('enter')
                                    # time.sleep(0.5)                    
                                    pyautogui.hotkey('ctrl','s')
                                    # close_form()   
                                    # time.sleep(1)
                                    pyautogui.hotkey('tab')            
                                    
                                    # ws['B'+str(rec)]='Completed'   
                                    ws['E'+str(rec)]='not default, inventory asset , based on rollup' 
                            # cost update, cost type 을 inventory asset 으로 변경 후 cost update
                            time.sleep(0.5)
                            # cost btn click
                            # cost_btn = pyautogui.locateOnScreen(resource_path(r'images\costs_btn.png'))
                            cost_btn = pyautogui.locateOnScreen(r'.\images\costs_btn.png')
                            if cost_btn is None: continue
                            cost_btn_center=pyautogui.center(cost_btn)
                            pyautogui.click(cost_btn_center)
                            # time.sleep(0.5)
                            pyautogui.write(cost_element1, interval=0.1)
                            # time.sleep(0.2)
                            pyautogui.hotkey('tab')
                            pyautogui.write(cost_element1, interval=0.1)
                            # time.sleep(0.2)      
                            # activity 칼럼 이동                  
                            pyautogui.press('tab')
                            # time.sleep(0.2)     
                            # basis 칼럼                   
                            pyautogui.press('tab')
                            # time.sleep(0.3)
                            pyautogui.write('Item', interval=0.1) 
                            # time.sleep(0.3)
                            pyautogui.hotkey('tab')
                            # time.sleep(0.2)
                            # Material cost 입력
                            pyautogui.write(str(cost), interval=0.1)
                            pyautogui.hotkey('tab')
                            # time.sleep(0.5)
                            pyautogui.hotkey('tab')
                            # material overhead
                            pyautogui.write(cost_element2, interval=0.1)
                            # time.sleep(0.2)
                            pyautogui.hotkey('tab')
                            # sub element
                            pyautogui.write(sub_element2, interval=0.1)
                            # time.sleep(0.4)
                            pyautogui.hotkey('tab')
                            # time.sleep(0.5)
                            pyautogui.press('tab') 
                            # time.sleep(0.5)
                            pyautogui.write('Total Value', interval=0.1)
                            # Material cost 입력         
                            pyautogui.hotkey('tab')
                            # time.sleep(0.3)
                            pyautogui.write(overhead, interval=0.1)
                            # time.sleep(0.2)
                            pyautogui.hotkey('tab')                        
                            # 저장
                            # time.sleep(0.3)
                            pyautogui.hotkey('ctrl','s')
                            # time.sleep(0.5)
                            pyautogui.hotkey('ctrl','f4')
                            # time.sleep(0.5)
                            pyautogui.hotkey('ctrl','f4')
                            # time.sleep(0.5)
                            
                            # print(item ,'peingind cost 있는 경우 Invetnroy Asset')
                            ws['E'+str(rec)]='Inventory Asset Completed'
                            print(item," Pending Cost Update 실행 시간 :", time.time() - start)      
            else:
                print(item," Item_cost1.prn  Not Found!")      
                
        # cost update process
        time.sleep(1)
        print('Frozen cost update')
        # std_menu = pyautogui.locateOnScreen(resource_path(r'images\standard.png'), grayscale= True, confidence=0.8)
        std_menu = pyautogui.locateOnScreen(r'.\images\standard.png')
        time.sleep(1)
        if std_menu is not None:
            pyautogui.doubleClick(pyautogui.center(std_menu))     
            # update menu
            # time.sleep(0.5)
            # # update cost menu 에서 down 으로 NOV Update Costs 메뉴로 이동
            # pyautogui.hotkey('down')
            time.sleep(1)
            rec=0
            for row_f in ws.iter_rows():
                rec += 1
                start = time.time()
                item_code = str(row_f[0].value)
                cost_type = str(row_f[1].value)
                # cost_type = 'Pending'
                if cost_type not in ['Pending', 'Annual'] : break
                
                x = dt.datetime.now()                
                # print(x.strftime('%Y%m%d%H%M'))
                cost_remark = 'KRK_'+ item_code+'_'+x.strftime('%Y%m%d%H%M')            
                time.sleep(1)    
                # update_menu = pyautogui.locateOnScreen(resource_path(r'images\nov_cost_update.png'), confidence=0.8)
                update_menu = pyautogui.locateOnScreen(r'.\images\nov_cost_update.png',confidence=0.8)
                #update_menu = pyautogui.locateOnScreen(r'.\images\nov_cost_update.png')
                time.sleep(1)
                if update_menu is not None:
                    update_menu_cen = pyautogui.center(update_menu)
                    # time.sleep(0.5)
                    pyautogui.doubleClick(update_menu_cen)
                    # time.sleep(0.5)
                    pyautogui.press('tab', presses=2, interval=0.3)
                    # time.sleep(0.3)
                    pyautogui.write(cost_type, interval=0.2)
                    pyautogui.hotkey('tab')
                    # time.sleep(0.3)
                    pyautogui.write(cost_remark, interval=0.1)
                    # time.sleep(0.3)
                    pyautogui.press('tab', presses=2, interval=0.3)
                    # time.sleep(0.3)
                    pyautogui.write(item_code, interval=0.1)
                    # time.sleep(0.3)
                    pyautogui.hotkey('tab')
                    # time.sleep(0.3)
                    pyautogui.hotkey('enter')
                    # time.sleep(0.3)
                    pyautogui.press('tab', presses=3, interval=0.3)
                    pyautogui.hotkey('enter')
                    # time.sleep(2)                    
                    # summit_btn = pyautogui.locateOnScreen(resource_path(r'images\summit_btn.png'), confidence=0.8)
                    summit_btn = pyautogui.locateOnScreen(r'.\images\summit_btn.png')
                    # time.sleep(0.5)
                    if summit_btn is not None:
                        # print('btn click')
                        pyautogui.click(pyautogui.center(summit_btn)) 
                        # time.sleep(1)
                        pyautogui.hotkey('tab')
                        # time.sleep(0.5)
                        pyautogui.hotkey('enter')
                        ws['F'+str(rec)]='Cost update run'
                        print(item_code," Cost Update 실행 시간 :", time.time() - start) 
                else:
                    print('NOV Update Costs menu 못 찾음')
                    pyautogui.screenshot(cost_remark)
                    break
                    
    print(" END 코드 실행 시간 :", time.time() - start1, rec) 
    wb.active = ws
    # wb.save(resource_path(r'data\Pending_JOB.xlsx')) 
    wb.save(r'.\data\Pending_JOB.xlsx')       
    pyautogui.alert('작업을 종료 하였습니다!!!!')     # type: ignore     
except Exception as e:
    # ws['E'+str(rec)]=e
    # wb.active = ws    
    # wb.save(resource_path(r'data\Pending_JOB.xlsx'))   
    print(e)    
   
    pyautogui.alert(text=e, button='OK')# type: ignore  
    time.sleep(10)    
   