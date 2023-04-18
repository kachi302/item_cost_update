# from ctypes.wintypes import SIZE
import sys
import os
import pyautogui
import time
from openpyxl import load_workbook

# print(fw, fw.title)
# if fw.isMaximized == False:
#     fw.maximize()
# def close_form():
#     time.sleep(1)  
#     closeri= pyautogui.locateOnScreen('close.png') 
#     if closeri != None:
#         closeric = pyautogui.center(closeri)
#         pyautogui.moveTo(closeric.x+(closeri.width/2)-10, closeric.y -(closeri.top/2)+20)   
#         pyautogui.click()
#     else:
#         print('close 화면 못 찾음')    
            
# def se_close_form():
#     time.sleep(1)  
#     mcloseri= pyautogui.locateOnScreen('close1.png') 
#     if mcloseri != None:
#         mcloseric = pyautogui.center(mcloseri)
#         pyautogui.moveTo(mcloseric.x+(mcloseri.width/2)-10, mcloseric.y -(mcloseri.top/2)+20)   
#         pyautogui.click()
#     else:
#         print('close1 화면 못 찾음')    
           
# def resource_path(relative_path):
#     try:
#         base_path = sys._MEIPASS  
#     except Exception:
#         base_path = os.path.abspath(".")

#     return os.path.join(base_path, relative_path)
# print(os.getcwd())
# print(os.path.dirname(os.path.realpath(__file__)))
running = True    

print('Annual Cost Reset Process, 3초후 시작 됩니다. 해당 화면을 Main 모니터가 아닌 다른 모니터로 이동 하세요.')
# time.sleep(3)
pyautogui.countdown(3)
    
# Inventory asset 에 대한 cost 값 0 처리를 위한 변수
cost_element1 = 'Material'
cost_element2 = 'Material Overhead'
sub_element2 = 'Freight'
# item_list =['10001285-001','10006047-001','10006048-001']
try:
    print('Start')
    print('Annual 초기화 대상 ITEM 자료 읽기....')
    # wb = load_workbook(resource_path(r'.\data\ANNUAL_list.xlsx'))
    wb = load_workbook(r'.\data\ANNUAL_list.xlsx')
    ws = wb['Sheet1']
    start1 = time.time()
    rec = 0
    time.sleep(3)
    job = pyautogui.confirm(text=' Annual Cost Reset 작업을 진행 하시겠습니까 ?', buttons =['OK','Cancel'])  # type: ignore
    
    if job == 'OK':
        for row in ws:
            if running == False:
                break            
            
            for cell in row:             # type: ignore
                        
                if len(cell.value) == 0 :
                    running = False
                    break            
                    
                print(cell.value)
                # 시작시간
                start = time.time()
                rec += 1
                time.sleep(0.5)
                
                # item_mp = pyautogui.locateOnScreen(resource_path(r'images\item_cost1.png'), confidence=0.7)     # type: ignore
                # cost menu click
                item_mp = pyautogui.locateOnScreen(r'.\images\item_cost1.png')                               
                print(rec)
                if item_mp != None:
                    # default 가 check 되어 있는 경우
                    # cost menu click
                    center=pyautogui.center(item_mp)
                    pyautogui.doubleClick(center)
                    time.sleep(0.8)
                    # pyautogui.write('10001285-001', interval=0.1)
                    # item code 입력
                    pyautogui.write(cell.value, interval=0.1)
                    time.sleep(0.8)
                    pyautogui.hotkey('tab')
                    time.sleep(0.8)
                    # cost type 입력
                    pyautogui.write('Annual', interval=0.1)  
                    time.sleep(0.8)
                    pyautogui.hotkey('tab')  
                    time.sleep(0.8)
                
                    # fi = pyautogui.locateOnScreen('find.png', confidence=0.7)
                    # find button click
                    # fi = pyautogui.locateOnScreen(resource_path(r'images\find.png'), confidence=0.7)
                    fi = pyautogui.locateOnScreen(r'.\images\find.png')
                    fic = pyautogui.center(fi) # type: ignore
                    pyautogui.click(fic)                    
                    time.sleep(1)      
                    # cost 화면 load     
                    # ri= pyautogui.locateOnScreen('basedroll1.png')
                    # ri= pyautogui.locateOnScreen(resource_path(r'images\basedroll1.png'))
                    # 1. 3가지 옵션이 다 있는 부분 check
                    ri= pyautogui.locateOnScreen(r'.\images\basedroll1.png')
                    if ri != None:
                        print(cell.value,' type : 3가지 옵션 다 있는 Item')
                        ric = pyautogui.center(ri) 
                        
                        # user defaul control uncheck
                        time.sleep(0.7)
                        pyautogui.click(ric.x-130, ric.y)
                        
                        # based on rollup unckeck
                        time.sleep(0.7)
                        pyautogui.click(ric.x+130, ric.y)
                        time.sleep(1)
                         # cost 지울지 메세지 나타나남, enter key 입력
                        pyautogui.hotkey('enter')
                        time.sleep(1)                                                 
                        # based on rollup ckeck
                        pyautogui.click(ric.x+130, ric.y)                       
                        time.sleep(1)  
                        
                        # user defaul control check
                        pyautogui.click(ric.x-130, ric.y)
                        time.sleep(1)
                        
                        pyautogui.hotkey('ctrl','s')
                        time.sleep(1)
                        pyautogui.hotkey('ctrl','f4')
                        # close_form()    
                        time.sleep(1)
                        ws['B'+str(rec)]='Completed'
                        print(cell.value,' 3 options, Completed')
                    else:
                        #user default control 가 unchecked 인 경우
                        time.sleep(1)
                        # ri_undefault = pyautogui.locateOnScreen('item_cost_not_default.png')
                        # ri_undefault = pyautogui.locateOnScreen(resource_path(r'images\item_cost_not_default.png'))
                        ri_undefault = pyautogui.locateOnScreen(r'.\images\item_cost_not_default.png')
                        
                        if ri_undefault != None:
                            print(cell.value,' type : User Default unckecked!')
                            ri_undefault_c = pyautogui.center(ri_undefault)                            
                            time.sleep(0.9)  
                            
                            pyautogui.click(ri_undefault_c.x+115, ri_undefault_c.y+10)
                            time.sleep(0.9)
                            
                            # cost 지울지 메세지 나타나남, enter key 입력
                            pyautogui.hotkey('enter')
                            time.sleep(0.9)
                            
                            pyautogui.click(ri_undefault_c.x+118, ri_undefault_c.y+10)
                            time.sleep(0.9)
                            pyautogui.hotkey('ctrl','s')
                            # close_form()   
                            time.sleep(1)
                            pyautogui.hotkey('ctrl','f4')               
                            time.sleep(1)
                            ws['B'+str(rec)]='Completed'
                        else:
                            time.sleep(1)
                            # inventory Asset 인 경우
                            # inv_asset = pyautogui.locateOnScreen(resource_path(r'images\inv_asset.png'))
                            print(cell.value,' type : Iventory Asset 만 있는 경우!')
                            inv_asset = pyautogui.locateOnScreen(r'.\images\inv_asset.png')
                            if  inv_asset != None :                                
                                time.sleep(0.9)
                                # cost_btn = pyautogui.locateOnScreen(resource_path(r'images\costs_btn.png'))
                                cost_btn = pyautogui.locateOnScreen(r'.\images\costs_btn.png')
                                cost_btn_center=pyautogui.center(cost_btn) # type: ignore
                                pyautogui.click(cost_btn_center)
                                time.sleep(0.9)                               
                                pyautogui.write(cost_element1, interval=0.1)
                                time.sleep(0.9)
                                pyautogui.hotkey('tab')
                                pyautogui.write(cost_element1, interval=0.1)
                                time.sleep(0.9)                        
                                pyautogui.press('tab', presses=3, interval=0.5) 
                                time.sleep(0.9)
                                # Material cost 입력
                                pyautogui.write(str(0), interval=0.1)
                                time.sleep(0.9)
                                pyautogui.hotkey('tab')
                                time.sleep(0.9)
                                pyautogui.hotkey('tab')
                                # material overhead
                                pyautogui.write(cost_element2, interval=0.1)
                                time.sleep(0.5)
                                pyautogui.hotkey('tab')
                                # sub element
                                time.sleep(0.5)
                                pyautogui.write(sub_element2, interval=0.1)
                                time.sleep(0.5)
                                pyautogui.hotkey('tab')
                                time.sleep(0.5)
                                pyautogui.press('tab', presses=2, interval=0.3) 
                                # Material Overhead rate 입력
                                time.sleep(0.9)                        
                                pyautogui.write(str(0), interval=0.1)
                                time.sleep(0.9)
                                pyautogui.hotkey('tab')                        
                                time.sleep(0.9)
                                pyautogui.hotkey('ctrl','s')
                                # close_form()   
                                time.sleep(0.9)
                                pyautogui.hotkey('ctrl','f4')
                                time.sleep(0.9)
                                pyautogui.hotkey('ctrl','f4')
                                print(cell.value ,'Invetnroy Asset')
                                ws['B'+str(rec)]='Inventory Asset Completed'
                                
                            else:        
                                print(cell.value ,'if 안 Annual 없음')                               
                                pyautogui.hotkey('ctrl','f4')
                                time.sleep(1)
                                pyautogui.hotkey('ctrl','f4')
                                time.sleep(1)
                                ws['B'+str(rec)]='Failed : if - not exists Annual'
                                
                else:
                    print(cell.value ,' 처음 Annual 없음')
                    pyautogui.hotkey('ctrl','f4')
                    time.sleep(1)
                    pyautogui.hotkey('ctrl','f4')
                    time.sleep(1)   
                    ws['B'+str(rec)]='First Screen Failed'
                    time.sleep(1)   
                    running = False         
                    break  
                time.sleep(1)
                print(cell.value," 코드 실행 시간 :", time.time() - start)  
                # pyautogui.hotkey('ctrl','f4')
                # time.sleep(1)
            
                        
    print(" END 코드 실행 시간 :", time.time() - start1, rec)   
    wb.active = ws
    wb.save(r'.\data\ANNUAL_list_JOB.xlsx')               
except Exception as e:
    wb.active = ws # type: ignore
    wb.save(r'.\data\ANNUAL_list_JOB.xlsx')    # type: ignore
    print(e)    
    time.sleep(10)