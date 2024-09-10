import pyautogui
import time
from openpyxl import load_workbook

def locate_and_click(image_path, confidence=0.8, click_offset=(0, 0), double_click=False):
    location = pyautogui.locateOnScreen(image_path, confidence=confidence)
    if location:
        center = pyautogui.center(location)
        pyautogui.click(center.x + click_offset[0], center.y + click_offset[1])
        if double_click:
            pyautogui.click(center.x + click_offset[0], center.y + click_offset[1])
        return True
    return False

def main():
    print('Start - 화면 크기 max 로')
    pyautogui.click(1000, 800)
    time.sleep(0.5)
    
    if not locate_and_click(r'.\images\screen_max.png', click_offset=(0, -20)):
        print('Screen Max not found!~')
        #return
    
    print('Max Screen')
    
    print('Excel File 읽기')
    wb = load_workbook(r'.\data\pending.xlsx', data_only=True)
    ws = wb['Sheet1']
    total_records = ws.max_row
    
    job = pyautogui.confirm(text=' Annual Cost Update 작업을 진행 하시겠습니끼?', buttons=['OK', 'Cancel'])
    if job != 'OK':
        return

    time.sleep(3)
    rec = 0
    start1 = time.time()
    for row in ws.iter_rows(min_row=1):  # Assuming the first row is header
        rec += 1
        start = time.time()
        item, cost_type, cost, overhead = [str(cell.value) for cell in row[:4]]
        print(item, cost_type, cost, overhead, total_records, rec)
        time.sleep(1)
        
        if locate_and_click(r'.\images\item_cost1.png', double_click=True):
            pyautogui.typewrite(item, interval=0.1)
            time.sleep(1)
            pyautogui.hotkey('tab')
            time.sleep(0.5)
            pyautogui.typewrite(cost_type, interval=0.1)
            time.sleep(0.5)
            pyautogui.hotkey('tab')
            time.sleep(0.5)
            
            if locate_and_click(r'.\images\find.png'):
                time.sleep(1)
                
                if locate_and_click(r'.\images\cost_type_not_found.png'):
                    print(item, '해당 cost type 이 없는 경우')
                    if locate_and_click(r'.\images\new.png'):
                        time.sleep(1)
                        pyautogui.typewrite(item, interval=0.1)
                        time.sleep(0.5)
                        pyautogui.hotkey('tab')
                        pyautogui.typewrite(cost_type, interval=0.1)
                        time.sleep(0.5)
                        
                        if locate_and_click(r'.\images\inv_asset.png', grayscale=True):       
                            
                            update_inventory_asset(cost, overhead)
                            ws['E' + str(rec)] = 'Inventory Asset Completed'
                            print(item, "코드 실행 시간 :", time.time() - start)
                        else:
                            print('Inventory Asset 이 아닌 경우')
                            ws['E' + str(rec)] = 'Inventory Non Asset item'
                
                else:
                    if locate_and_click(r'.\images\inv_asset.png'):
                        print(item, ' Invetnroy Asset')
                        update_inventory_asset(cost, overhead)
                        ws['E' + str(rec)] = 'Inventory Asset Completed'
                        print(item, "Annual cost update 실행 시간 :", time.time() - start)
                    else:
                        print(item, ' Others Invetnroy Asset')
                        handle_default_values(item)
                        update_inventory_asset(cost, overhead)
                        ws['E' + str(rec)] = 'Inventory Asset Completed'
                        print(item, "Annual Cost Update 실행 시간 :", time.time() - start)
            
            time.sleep(0.5)
            pyautogui.hotkey('ctrl', 'f4')
            time.sleep(0.5)
            pyautogui.hotkey('ctrl', 'f4')
            time.sleep(0.5)

    wb.save(r'.\data\pending_updated.xlsx')
    print(" END 코드 실행 시간 :", time.time() - start1, rec)  
    pyautogui.alert('작업을 종료 하였습니다!!!!')  
    
def update_inventory_asset(cost, overhead):
    cost_btn = pyautogui.locateOnScreen(r'.\images\costs_btn.png')
    cost_btn_center=pyautogui.center(cost_btn)# type: ignore
    pyautogui.click(cost_btn_center)
    time.sleep(0.5)
                            
    pyautogui.hotkey('ctrl', 'l')
    time.sleep(0.5)
    pyautogui.hotkey('enter')
    time.sleep(0.5)
    pyautogui.hotkey('ctrl', 'l')
    time.sleep(0.5)
    pyautogui.hotkey('tab')
    time.sleep(0.5)
    pyautogui.hotkey('tab')
    time.sleep(0.5)
    pyautogui.typewrite(str(cost), interval=0.1)
    time.sleep(0.5)
    pyautogui.hotkey('tab')
    time.sleep(0.7)
    pyautogui.hotkey('tab')
    time.sleep(0.8)
    pyautogui.hotkey('ctrl', 'l')
    time.sleep(0.5)
    pyautogui.press('down')
    time.sleep(0.5)
    pyautogui.hotkey('enter')
    time.sleep(0.5)
    pyautogui.hotkey('ctrl', 'l')
    time.sleep(0.5)
    pyautogui.press('down')
    time.sleep(0.5)
    pyautogui.hotkey('enter')
    time.sleep(0.7)
    pyautogui.hotkey('tab')
    time.sleep(0.7)
    pyautogui.hotkey('ctrl', 'l')
    time.sleep(0.5)
    pyautogui.press('down', presses=5)
    time.sleep(0.5)
    pyautogui.hotkey('enter')
    time.sleep(0.5)
    pyautogui.typewrite(str(overhead), interval=0.1)
    time.sleep(0.8)
    pyautogui.hotkey('ctrl', 's')
    time.sleep(0.5)

def handle_default_values(item):
    if locate_and_click(r'.\images\basedroll1.png'):
        print(item, 'type : 3가지 다 있는 경우,default')
        pyautogui.click(pyautogui.center(location).x - 130, pyautogui.center(location).y)
        time.sleep(0.5)
        pyautogui.click(pyautogui.center(location).x + 130, pyautogui.center(location).y)
        time.sleep(1)
        pyautogui.hotkey('enter')
        time.sleep(1)
        pyautogui.hotkey('ctrl', 's')
        time.sleep(1)
        pyautogui.hotkey('tab')
    elif locate_and_click(r'.\images\item_cost_not_default.png'):
        print(item, 'type : Inventory, Based on Rollup')
        pyautogui.click(pyautogui.center(location).x + 115, pyautogui.center(location).y + 10)
        time.sleep(1)
        pyautogui.hotkey('enter')
        time.sleep(0.5)
        pyautogui.hotkey('ctrl', 's')
        time.sleep(1)
        pyautogui.hotkey('tab')

if __name__ == '__main__':
    main()