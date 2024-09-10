import sys
import os

import pyautogui as pag
import time
import datetime as dt
from openpyxl import load_workbook
import pymsgbox as pmg
from pyscreeze import locateAllOnScreen

running = True
print('Item Mfg Group 추가...3초후 시작됩니다. 해당 화면을 다른 모니터로 이동 하세요.')
pag.countdown(3)
# item_code = ''
# planner_code = ''18669305-004
try:
    print('Start')
    # pag.click(1000, 800)
    wb = load_workbook(r'D:\python_dev\item_cost_update\mfgAdd.xlsx', data_only= True)
    ws = wb['Sheet1']
    start1 = time.time()
    rec = 0
    time.sleep(1)
    dest = "D:\\python_dev\\item_cost_update\\images\\"
    # job = pag..confirm('Item Planner 변경 작업을 진행 하시겠습니끼?', buttons= ['OK','Cancel']) # type: ignore    
    job = pmg.confirm('Item MFG 추가 작업을 진행 하시겠습니끼?', buttons= ['OK','Cancel']) 
    if job == 'OK':
        print(time.strftime('%Y.%m%d - %H:%M:%S'))
        time.sleep(3)
        for row in ws.iter_rows():
            rec += 1
            item = str(row[0].value)
            time.sleep(0.3)
            item_len = len(item)
            mfg_group = str(row[1].value)
            mfg_value = str(row[2].value)
            planning = str(row[3].value)
            planning_value = str(row[4].value)
            print('Rec : %d , item = %s' % (rec,item))
            if item =='E':
                print('End....')
                break
            if item_len <= 11:
                print('Rec : %d , item = %s, Categories : %s' % (rec,item, 'item error', mfg_value))
                break
            # item 차기
            pag.hotkey('f11')
            time.sleep(1)
            pag.write(item, interval=0.03)
            time.sleep(0.5)
            pag.hotkey('ctrl','f11')
            time.sleep(1)
            # tool button 찾기
            tools_png = pag.locateOnScreen(r'D:\python_dev\item_cost_update\images\ITEM_TOOLS.png', confidence=0.8)
            if tools_png is not None:
                pag.click(pag.center(tools_png))
                time.sleep(0.5)
                pag.press('down')        
                time.sleep(0.3)  
                # tools -- categories
                # print('down after')
                category_png = pag.locateOnScreen(r'D:\python_dev\item_cost_update\images\Categories.png', confidence=0.8)     
                if category_png is not None:
                    pag.click(pag.center(category_png))
                
                    time.sleep(2)
                    pag.hotkey('ctrl','down')
                    # time.sleep(1)
                    pag.write(mfg_group, interval=0.03)
                    time.sleep(0.3)
                    # pag.press('enter')
                    pag.press('tab')
                    time.sleep(0.3)
                    pag.write(mfg_value, interval=0.03)               
                    time.sleep(0.5)
                    pag.hotkey('ctrl','s')
                    time.sleep(7)
                    # planning group 추가
                    pag.press('tab')
                    time.sleep(2)
                    pag.hotkey('f11')
                    time.sleep(0.2)
                    pag.write(planning,interval=0.03)
                    pag.press('tab', presses=3)
                    pag.press('enter')
                    time.sleep(0.5)
                    pag.press('tab')
                    pag.write(planning_value, interval=0.03)
                    pag.hotkey('ctrl','s')
                    time.sleep(7)
                    
                    
                    # pag.hotkey('ctrl','f4')
                    close_png = pag.locateOnScreen(r'D:\python_dev\item_cost_update\images\Close_page.png', confidence=0.8)
                    if close_png is not None:
                        pag.click(pag.center(close_png))
                    time.sleep(1)
            else:
                print('Rec : %d , item = %s, %s' % (rec,item, 'tools_png error'))
                pag.screenshot(dest+"tools_png.jpg")
                time.sleep(1)
                break   
            
    print(time.strftime('%Y.%m%d - %H:%M:%S'))        
    pmg.alert('작업을 종료합니다.')
    print('작업을 종료합니다.')
except Exception as e:
    # wb.active = ws # type: ignore
    # wb.save(r'.\Planner_job.xlsx')    # type: ignore
    print(e)    
    time.sleep(10)       