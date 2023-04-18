import sys
import os
import pyautogui
import time
import datetime as dt
from openpyxl import load_workbook
print('Frozen Cost Update Process, 5초후 시작 됩니다. 해당 화면을 다른 모니터로 이동 하세요.')
time.sleep(5)
rec=0
try:
    print('Excel File 읽기')           
    wb = load_workbook(r'.\data\pending.xlsx', data_only= True)
    ws = wb['Sheet1']
    # 전체 작업 시간
    start1 = time.time()
    job = pyautogui.confirm(text=' Frozen Cost Update 작업을 진행 하시겠습니끼?', buttons =['OK','Cancel'])# type: ignore  
    if job == 'OK':
        
        time.sleep(1)
        print('Frozen cost update')
        # std_menu = pyautogui.locateOnScreen(resource_path(r'images\standard.png'), grayscale= True, confidence=0.8)
        std_menu = pyautogui.locateOnScreen(r'.\images\standard.png')
        if std_menu is not None:
            pyautogui.doubleClick(pyautogui.center(std_menu))     
            # update menu
            time.sleep(0.5)
            # update cost menu 에서 down 으로 NOV Update Costs 메뉴로 이동
            # pyautogui.hotkey('down')
            # time.sleep(0.2)
            rec=0
            for row_f in ws.iter_rows():
                rec += 1
                start = time.time()
                item_code = str(row_f[0].value)
                cost_type = 'Pending'
                x = dt.datetime.now()                
                # print(x.strftime('%Y%m%d%H%M'))
                cost_remark = 'KRK_'+ item_code+'_'+ x.strftime('%Y%m%d%H%M')            
                time.sleep(1)    
                # update_menu = pyautogui.locateOnScreen(resource_path(r'images\nov_cost_update.png'), confidence=0.8)
                update_menu = pyautogui.locateOnScreen(r'.\images\nov_cost_update.png')
                if update_menu is not None:
                    update_menu_cen = pyautogui.center(update_menu)
                    pyautogui.doubleClick(update_menu_cen)
                    time.sleep(0.5)
                    pyautogui.press('tab', presses=2, interval=0.3)
                    time.sleep(0.3)
                    pyautogui.write(cost_type, interval=0.2)
                    pyautogui.hotkey('tab')
                    time.sleep(0.3)
                    pyautogui.write(cost_remark, interval=0.1)
                    time.sleep(0.3)
                    pyautogui.press('tab', presses=2, interval=0.3)
                    time.sleep(0.3)
                    pyautogui.write(item_code, interval=0.1)
                    time.sleep(0.3)
                    pyautogui.hotkey('tab')
                    time.sleep(0.3)
                    pyautogui.hotkey('enter')
                    time.sleep(0.3)
                    pyautogui.press('tab', presses=3, interval=0.3)
                    pyautogui.hotkey('enter')
                    time.sleep(2)
                    # summit_btn = pyautogui.locateOnScreen(resource_path(r'images\summit_btn.png'), confidence=0.8)
                    summit_btn = pyautogui.locateOnScreen(r'.\images\summit_btn.png')
                    if summit_btn is not None:
                        # print('btn click')
                        pyautogui.click(pyautogui.center(summit_btn)) 
                        time.sleep(1)
                        pyautogui.hotkey('tab')
                        time.sleep(0.5)
                        pyautogui.hotkey('enter')
                        ws['F'+str(rec)]='Cost update run'
                        print(item_code," Cost Update 실행 시간 :", time.time() - start) 
                else:
                    print('NOV Update Costs menu 못 찾음')
                    break
                
    print(" END 코드 실행 시간 :", time.time() - start1, rec)   
    wb.active = ws
    # wb.save(resource_path(r'data\Pending_JOB.xlsx')) 
    wb.save(r'.\data\Frozen_JOB.xlsx')   
    pyautogui.alert('작업을 종료 하였습니다.!')   # type: ignore            
except Exception as e:
    # ws['E'+str(rec)]=e
    # wb.active = ws    
    # wb.save(resource_path(r'data\Pending_JOB.xlsx'))   
    print(e)    
    pyautogui.alert(text=e, button='OK')# type: ignore  
    time.sleep(10)                    